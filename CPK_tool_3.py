import pandas as pd
import numpy as np
import matplotlib.pyplot as plt

# ===============================
# CONVERSION DE COLUMNAS (CRÍTICO)
# ===============================
def convertir_columnas(df):
    for col in ["W", "U", "V"]:
        df[col] = df[col].astype(str)
        df[col] = df[col].str.replace(",", ".", regex=False)
        df[col] = df[col].str.strip()
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

# ===============================
# CALCULO CPK
# ===============================
def calcular_cpk(df):

    resultados = []

    grupos = df.groupby("O")

    for key, grupo in grupos:

        valores = grupo["W"].dropna().values

        if len(valores) < 2:
            continue

        avg = np.mean(valores)
        stdev = np.std(valores, ddof=1)

        lsl = grupo["U"].iloc[0]
        usl = grupo["V"].iloc[0]

        if lsl == usl:
            cpl = cpu = cpk = None
            comentario = "LSL = USL (No SPC)"

        elif stdev == 0:
            cpl = cpu = cpk = None
            comentario = "Stdev = 0 (Proceso constante)"

        else:
            cpl = (avg - lsl) / (3 * stdev)
            cpu = (usl - avg) / (3 * stdev)
            cpk = min(cpl, cpu)
            comentario = ""

        resultados.append({
            "Test Step": key,
            "LSL": lsl,
            "USL": usl,
            "Average": round(avg, 6),
            "Stdev.S": round(stdev, 6),
            "CPL": None if cpl is None else round(cpl, 4),
            "CPU": None if cpu is None else round(cpu, 4),
            "CPK": None if cpk is None else round(cpk, 4),
            "Comentario": comentario
        })

    return pd.DataFrame(resultados)

# ===============================
# EXPORTAR A EXCEL CON COLORES
# ===============================
def exportar_excel(df_res, ruta_salida):

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:

        df_res.to_excel(writer, index=False, sheet_name="CPK")

        wb = writer.book
        ws = writer.sheets["CPK"]

        from openpyxl.styles import PatternFill

        red = PatternFill(start_color="FF0000", fill_type="solid")
        yellow = PatternFill(start_color="FFFF00", fill_type="solid")
        green = PatternFill(start_color="00B050", fill_type="solid")
        gray = PatternFill(start_color="C0C0C0", fill_type="solid")

        # Columna CPK = H
        for row in range(2, len(df_res) + 2):

            cell = ws[f"H{row}"]

            if cell.value is None:
                cell.fill = gray
            elif cell.value < 1:
                cell.fill = red
            elif cell.value < 1.33:
                cell.fill = yellow
            else:
                cell.fill = green

    print(f"✅ Archivo generado: {ruta_salida}")

# ===============================
# GRAFICA DOTPLOT
# ===============================
def graficar_test_step(df, test_step):

    subset = df[df["O"] == test_step]

    if len(subset) < 2:
        print("No hay suficientes datos")
        return

    y = subset["W"].values
    x = np.arange(1, len(y)+1)

    lsl = subset["U"].iloc[0]
    usl = subset["V"].iloc[0]

    plt.figure(figsize=(12,6))
    plt.gca().set_facecolor('#f2f2f2')

    # Puntos
    plt.scatter(x, y, color='red', edgecolors='blue', s=60, label="Measurement")

    # Líneas
    plt.axhline(y=lsl, color='blue', linewidth=2, label="LSL")
    plt.axhline(y=usl, color='blue', linestyle='--', linewidth=2, label="USL")

    # Labels
    plt.title(test_step, fontsize=16)
    plt.xlabel("Sample Index")
    plt.ylabel("Measurement")

    plt.grid(True)
    plt.legend()

    plt.show()

# ===============================
# MAIN
# ===============================
if __name__ == "__main__":

    ruta = input("Ruta del archivo Excel: ")

    try:
        df = pd.read_excel(ruta, sheet_name="Sheet1")

        # LIMPIEZA
        df = convertir_columnas(df)
        df["O"] = df["O"].astype(str).str.strip()
        df = df.dropna(subset=["O", "W"])

        print("\nCalculando CPK...\n")

        df_res = calcular_cpk(df)

        print(df_res.head())

        # EXPORTAR
        exportar_excel(df_res, "CPK_Dashboard.xlsx")

        # ===============================
        # PEDIR TEST STEP PARA GRAFICAR
        # ===============================
        print("\nLista de Test Steps:")
        print(df_res["Test Step"].tolist())

        test = input("\nIngresa Test Step a graficar (o ENTER para salir): ")

        if test != "":
            graficar_test_step(df, test)

    except Exception as e:
        print(f"Error: {e}")