import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ===============================
# FUNCION CONVERSION (CRITICA)
# ===============================
def convertir_columna(col):
    col = col.astype(str)
    col = col.str.replace(",", ".", regex=False)
    col = col.str.strip()
    return pd.to_numeric(col, errors="coerce")

# ===============================
# MAIN
# ===============================
archivo = input("Ingresa ruta del archivo Excel: ")

# Leer archivo
df = pd.read_excel(archivo, sheet_name="Sheet1")

# ===============================
# VALIDAR COLUMNAS
# ===============================
for col in ["O","U","V","W"]:
    if col not in df.columns:
        print(f"Falta columna {col}")
        exit()

# ===============================
# LIMPIEZA
# ===============================
df["O"] = df["O"].astype(str).str.strip()
df["W"] = convertir_columna(df["W"])
df["U"] = convertir_columna(df["U"])
df["V"] = convertir_columna(df["V"])

df = df.dropna(subset=["O","W"])

# ===============================
# CALCULO CPK
# ===============================
resultados = []

for step, grupo in df.groupby("O"):

    valores = grupo["W"].values

    if len(valores) < 2:
        continue

    avg = np.mean(valores)
    std = np.std(valores, ddof=1)

    lsl = grupo["U"].iloc[0]
    usl = grupo["V"].iloc[0]

    if lsl == usl:
        cpl = cpu = cpk = None
        comentario = "LSL = USL"

    elif std == 0:
        cpl = cpu = cpk = None
        comentario = "Stdev = 0"

    else:
        cpl = (avg - lsl) / (3 * std)
        cpu = (usl - avg) / (3 * std)
        cpk = min(cpl, cpu)
        comentario = ""

    resultados.append([
        step, lsl, usl,
        avg, std,
        cpl, cpu, cpk,
        comentario
    ])

# Crear DataFrame
cols = ["Test Step","LSL","USL","Average","Stdev","CPL","CPU","CPK","Comentario"]
df_res = pd.DataFrame(resultados, columns=cols)

# ===============================
# ESCRIBIR EXCEL
# ===============================
with pd.ExcelWriter(archivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_res.to_excel(writer, sheet_name="CPK_Summary", index=False)

# ===============================
# FORMATO COLORES
# ===============================
wb = load_workbook(archivo)
ws = wb["CPK_Summary"]

rojo = PatternFill(start_color="FF0000", fill_type="solid")
verde = PatternFill(start_color="00B050", fill_type="solid")
gris = PatternFill(start_color="C0C0C0", fill_type="solid")

for i in range(2, ws.max_row + 1):

    cpk_cell = ws[f"H{i}"]

    if cpk_cell.value is None:
        cpk_cell.fill = gris
    elif cpk_cell.value < 1.33:
        cpk_cell.fill = rojo
    else:
        cpk_cell.fill = verde

wb.save(archivo)

print("✅ CPK calculado y guardado en Excel")

# ===============================
# GRAFICA
# ===============================
resp = input("¿Deseas graficar un Test Step? (y/n): ")

if resp.lower() == "y":

    print("\nTest Steps disponibles:")
    print(df_res["Test Step"].tolist())

    step_sel = input("Ingresa el Test Step: ")

    subset = df[df["O"] == step_sel]

    if len(subset) < 2:
        print("No hay datos suficientes")
    else:

        y = subset["W"].values
        x = np.arange(1, len(y)+1)

        lsl = subset["U"].iloc[0]
        usl = subset["V"].iloc[0]

        plt.figure(figsize=(12,6))
        plt.gca().set_facecolor("#f2f2f2")

        plt.scatter(x, y, color='red', edgecolors='blue', s=60)

        plt.axhline(y=lsl, color='blue', linewidth=2, label="LSL")
        plt.axhline(y=usl, color='blue', linestyle='--', linewidth=2, label="USL")

        plt.title(step_sel)
        plt.xlabel("Sample Index")
        plt.ylabel("Measurement")

        plt.grid(True)
        plt.legend()

        plt.show()